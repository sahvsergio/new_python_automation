"""
modules to be used
csv (https://docs.python.org/2/library/csv.html)
openpyxl (https://pypi.python.org/pypi/openpyxl)
XlsxWriter (https://pypi.python.org/pypi/XlsxWriter)
"""

import xlsxwriter
import csv
"""
fh = open("iris_2b6b1f4d-1c15-4a87-967d-06f0a88fbe13.csv", 'rt') 
try: 
    reader = csv.reader(fh) 
    print ("Data from the CSV:", list(reader)) 
except Exception as e: 
    print ("Exception is:", e)
finally: 
    fh.close()

print( '*'*100)
"""
"""
try: 
    reader = csv.reader(open("mylist.csv", 'rU'), 
                                dialect=csv.excel_tab) 
    print ("Data from the CSV:") 
    d = list(reader) 
    print ("\n").join("%-20s %s"%(d[i],d[i+len(d)/2]) for i in range(len(d)/2)) 
 
except Exception as e: 
    print ("Exception is:", e) 
finally: 
    fh.close()
    """
print('*'*100)
"""

try: 
    reader = csv.reader(open("iris_2b6b1f4d-1c15-4a87-967d-06f0a88fbe13.csv", ), dialect=csv.excel_tab) 
    print ("Data from the CSV:" )
    d = list(reader) 
    print( "\n").join("%-20s %s"%(d[i],d[i+len(d)/2]) for i in range(len(d)/2)) 
 
except Exception as e: 
    print ("Exception is:", e) 
finally: 
    fh.close()

"""
print('*'*100)
print('')

"""
fh = open("iris_2b6b1f4d-1c15-4a87-967d-06f0a88fbe13.csv", 'rt') 
reader = csv.reader(fh) 
data = list(reader) 
print ("Data cells from CSV:") 
print (data[0][1], data[1][1] )
print (data[0][2], data[1][2] )
print (data[0][3], data[1][3] )
"""

"""In the preceding code example, we print a part of the data from the CSV file. If you realize, a CSV file can also be read as a 2D list in Python with the first index as row and the second index as column. Here we print the second, third, and fourth columns from row1 and row2:"""
print('*'*100)
"""#reading from a dictionary

f = open("iris_2b6b1f4d-1c15-4a87-967d-06f0a88fbe13.csv", 'rt') 
print ("File Contents") 
try: 
    reader = csv.DictReader(f) 
    for row in reader: 
        print (row['sepal.length'], row['sepal.width'], row['petal.length'],row['petal.width'], row['variety'] )
finally:
    f.close()
"""

"sepal.length", "sepal.width", "petal.length", "petal.width", "variety"

# helper methods  also known  as reader objects
"""
import csv 
f = open("iris_2b6b1f4d-1c15-4a87-967d-06f0a88fbe13.csv", 'rt') 
reader = csv.DictReader(f) 
print ("Columns in CSV file:", reader.fieldnames) 
print ("Dialect used in CSV file:", reader.dialect) 
print ("Current line number in CSV file:", reader.line_num) 
print ("Moving the reader to next line with reader.next()") 
#reader.next()  changged to next(reader)
next(reader)
print ("Reading line number:", reader.line_num)
f.close()
        
#names: Gives list of column names
#dialect: CSV file format (we 'll read more about it)
#line_num: Current line number being read
#next():Takes you to the next line
"""
# writing data  into csv files

names = ["John", 'Eve', 'Fate', 'Jadon']
grades = ['C', 'A+', 'A', 'B-']
f = open('newlist.csv', 'wt')  # creating the file
try:
    writer = csv.writer(f)  # opening the file
    writer.writerow(('Names', 'Grades'))  # creating headers
    for i in range(4):
        # separated rows by commas
        writer.writerow((names[i], grades[i]))

finally:
    f.close()

f = open("write.csv", 'wt')
csvWriter = csv.writer(f, delimiter='\t',
                       lineterminator='\n\n')
csvWriter.writerow(['abc', 'pqr', 'xyz'])
csvWriter.writerow(['123', '456', '789'])
f.close()
# print  available dialects
print("Available Dialects:", csv.list_dialects())

# register our own dialects- for example with our own delimiter
# register the dialect
"""
names=["John", 'Eve','Fate','Jadon']
grades=['C','A+','A','B-']
csv.register_dialect('pipes', delimiter='-')
#open file
with open('pipes.csv', 'r') as f:
   
    reader = csv.reader(f, dialect='pipes')
    for row in reader:        
        print(row)
"""
# Managing employee information in an automated way

# open the original file to read it
"""
f = open("mylist.csv", 'rt')
#open the second file for writing
fw = open("CA_Employees.csv", 'wt')


try:
    #passing the first file to the dict reader
    reader=csv.DictReader(f)
    #creating a csv.writer object with file to write on the newer file
    csv_Writer=csv.writer(fw)
    #loop throught the first document
    for row in reader:
        #skipping the 1st row which contains the header
        if reader.line_num==1:
            continue
        #identify the row with state ==CA
        
        if row['state']=='CA':
            #write both the email and phone number
            csv.writerow([row['email'],row['phone']])
finally:
    #close both files
    f.close()
    fw.close()
                 
"""
"""
print('*'*50, 'Reading Excel Sheets','*'*14)   
#Reading Excel sheets
import openpyxl #-->the module that will be used
#load the  excel file
workbook=openpyxl.load_workbook('SalesData.xlsx')
#get the sheet_names
print('Workbook Object:',workbook.get_sheet_names())
#accessing  a specific sheet inside the worksheet
people=workbook.get_sheet_by_name('Sheet1')
print('Sheet 1  sheet Object:', people)

#reading the cell objects
print("First People Object:",people['A1'])
print("Other Cell Object:",people.cell(row=3, column=2))

#Getting Values of the cell 
print('First Name:',people['B2'].value,people['C2'].value)
"""
# writing excel

# creating a worksheet
workbook = xlsxwriter.Workbook('add_sheet.xlsx')
# create a sheet
worksheet = workbook.add_worksheet(name='New Sheet2')

workbook.close()  # if you close without adding a  new sheet it creates with a basic sheet


# writing operations on excel worksheets
workbook = xlsxwriter.Workbook('Expenses01.xlsx')
worksheet = workbook.add_worksheet()
# create a tuple of lists  or any data structure to hold the data to be entered
expenses = (
    ['Rent', 1000],
    ['Gas', 100],
    ['Food', 300],
    ['Gym', 50]
)
# checking the types as textbook stated that this was a dictionary
print(type(expenses))  # tuple
print(type(expenses[0]))  # list
# initializers for rows and cols
row = 0
col = 0
# iterating over  the tuple items, to add items and cost
for item, cost in expenses:
    # tell what row and col to write to and what to write
    worksheet.write(row, col, item)
    # addig  a new row
    worksheet.write(row, col+1, cost)
    row += 1

worksheet.write(row, 0, 'Total')
worksheet.write(row, 1, '=SUM(B1:B4)')

workbook.close()
# adding format to cell
"""format=workbook.add_format()
format.set_bold
format.set_font_color('green')
"""


workbook = xlsxwriter.Workbook('cell_format.xlsx')
worksheet = workbook.add_worksheet()
expenses = (
    ['Rent', 1000],
    ['Gas',   100],
    ['Food',  300],
    ['Gym',    50],
)

row = 0
col = 0
for item, cost in (expenses):
    worksheet.write(row, col, item)
    worksheet.write(row, col + 1, cost)
    row += 1

format1 = workbook.add_format({'bg_color': 'blue',
                               'font_color': 'red'})

worksheet.conditional_format('B1:KB5',
                             {'type': 'cell',
                              'criteria': '>=',
                              'value': 150,
                              'format': format1}
                             )
workbook.close()

# Playing with Excel formulae
workbook = xlsxwriter.Workbook('formula.xlsx')
worksheet = workbook.add_worksheet()
# add a list numbers and store the sum to Cell A1

worksheet.write_formula('A1', 'SUM(1,2,3)')

# building charts
workbook = xlsxwriter.Workbook('chartline.xlsx')
worksheet = workbook.add_worksheet()
data = [10, 40, 50, 20, 10, 50]
worksheet.write_column('A1', data)
chart = workbook.add_chart({'type': 'line'})

# this takes the range in the excel sheet where data is located
chart.add_series({'values': '=Sheet1!A$1:$A$6'})

# takes the cell  name and chart object as arugments
worksheet.insert_chart('C1', chart)
workbook.close()


workbook = xlsxwriter.Workbook('chart_column.xlsx')
worksheet = workbook.add_worksheet()
chart = workbook.add_chart({'type': 'column'})
data = [
    ['Year', '2013', '2014', '2015'],
    ['Revenue', 100, 120, 125],
    ['COGS', 80, 90, 70],
    ['Profit', 20, 30, 55]
]

worksheet.write_row('A1', data[0])
worksheet.write_row('A2', data[1])
worksheet.write_row('A3', data[2])
worksheet.write_row('A4', data[3])

chart.add_series({'values': '=Sheet1!$B2:$B$4', 'name': '2013'})
chart.add_series({'values': '=Sheet1!$C2:$C$4', 'name': '2014'})
chart.add_series({'values': '=Sheet1!$D2:$D$4', 'name': '2015'})
worksheet.insert_chart('G1',chart)
worksheet.write(5,0,'% Gain')
worksheet.write(5,1,'=(B4/B2)*100')
worksheet.write(5,2,'=(C4/C2)*100')
worksheet.write(5,3,'=(D4/D2)*100')
workbook.close()


